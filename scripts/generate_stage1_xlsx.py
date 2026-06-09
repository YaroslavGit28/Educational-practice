"""Генерация xlsx-файлов для сдачи этапа 1 УП.03."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
STAGE1 = ROOT / "сдача" / "УП03_Этап1_Ярченко"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)


def save_stage1_xlsx():
    STAGE1.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Состав проекта"
    ws.append(["Файл/папка", "Назначение", "Статус"])
    for row in [
        ("apps/", "Django-приложения (users, catalog, cart, orders, processing)", "OK"),
        ("magazin/", "Настройки проекта, health, reports", "OK"),
        ("templates/", "HTML-шаблоны интерфейса", "OK"),
        ("static/css/", "Стили Bootstrap", "OK"),
        ("tests/", "Автотесты Django TestCase", "OK"),
        ("scripts/", "BAT-скрипты запуска и проверки", "OK"),
        ("docs/", "Отчёты и планы тестирования", "OK"),
        ("requirements.txt", "Зависимости Python", "OK"),
        ("manage.py", "Точка входа Django", "OK"),
        ("db.sqlite3", "База данных SQLite", "OK"),
        ("README.md", "Описание проекта", "OK"),
        (".env.example", "Пример настроек", "OK"),
        ("Dockerfile", "Контейнеризация", "OK"),
        ("docker-compose.yml", "Локальный Docker", "OK"),
        ("docker-compose.prod.yml", "Production/demo compose", "OK"),
        (".github/workflows/ci.yml", "CI pipeline", "OK"),
    ]:
        ws.append(list(row))
    style_header(ws)
    wb.save(STAGE1 / "02_Состав_проекта.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Конфигурация"
    ws2.append(["Переменная", "Пример", "Назначение", "Обязательна"])
    for row in [
        ("SECRET_KEY", "change_me_in_real_project", "Секретный ключ Django", "Да"),
        ("DEBUG", "true / false", "Режим отладки", "Да"),
        ("ALLOWED_HOSTS", "localhost,127.0.0.1", "Разрешённые хосты", "Да"),
        ("APP_PORT", "8000", "Порт приложения", "Нет"),
        ("DATABASE_URL", "sqlite:///db.sqlite3", "Подключение к БД", "Да"),
        ("LOG_LEVEL", "INFO", "Уровень логирования", "Нет"),
        ("APP_ENV", "development / demo / production", "Режим окружения", "Нет"),
        ("DEPLOY_TARGET", "docker_local", "Способ развертывания", "Нет"),
    ]:
        ws2.append(list(row))
    style_header(ws2)
    wb2.save(STAGE1 / "04_Таблица_конфигурации.xlsx")

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Журнал проблем"
    ws3.append(["ID", "Проблема", "Как проявляется", "Причина", "Решение", "Статус"])
    for row in [
        ("P-01", "Проект не запускается", "ModuleNotFoundError: users", "apps/ не в PYTHONPATH", "sys.path.insert в manage.py", "Исправлено"),
        ("P-02", "Нет миграций", "no such table", "Миграции не созданы", "makemigrations + migrate", "Исправлено"),
        ("P-03", "Нет шаблонов", "TemplateDoesNotExist", "13 шаблонов отсутствовали", "Созданы все шаблоны", "Исправлено"),
        ("P-04", "Нет .env.example", "Секреты в коде", "Не настроено", "Создан .env.example", "Исправлено"),
        ("P-05", "Авторизация по email", "Неверный логин", "Нет EmailBackend", "users/backends.py", "Исправлено"),
        ("P-06", "Нет README", "Непонятно как запустить", "Документация отсутствует", "README + INSTALL", "Исправлено"),
        ("P-07", "Нет wsgi.py", "Ошибка deploy", "Файл отсутствовал", "Создан magazin/wsgi.py", "Исправлено"),
    ]:
        ws3.append(list(row))
    style_header(ws3)
    wb3.save(STAGE1 / "05_Журнал_проблем.xlsx")

    print(f"XLSX created: {STAGE1}")


if __name__ == "__main__":
    save_stage1_xlsx()
