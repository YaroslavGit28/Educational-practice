from datetime import datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from orders.models import Order, OrderItem
from catalog.models import Product


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('users:login')
        if request.user.role not in ('admin', 'manager') and not request.user.is_superuser:
            messages.error(request, 'Нет доступа к отчётам')
            return redirect('catalog:index')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@admin_required
def reports_index(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders = Order.objects.exclude(status='cancelled')
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    orders_count = orders.count()

    top_products = (
        OrderItem.objects.filter(order__in=orders)
        .values('product__name')
        .annotate(qty=Sum('quantity'), revenue=Sum('price'))
        .order_by('-qty')[:10]
    )

    status_stats = orders.values('status').annotate(count=Count('id'))

    return render(request, 'reports/index.html', {
        'total_revenue': total_revenue,
        'orders_count': orders_count,
        'top_products': top_products,
        'status_stats': status_stats,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
@admin_required
def export_sales_excel(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders = Order.objects.exclude(status='cancelled').select_related('user')
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Продажи'

    headers = ['№ заказа', 'Дата', 'Покупатель', 'Статус', 'Сумма', 'Телефон']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for order in orders.order_by('-created_at'):
        ws.append([
            order.id,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            order.user.full_name,
            order.get_status_display(),
            float(order.total_amount),
            order.phone,
        ])

    ws2 = wb.create_sheet('Товары')
    ws2.append(['Товар', 'Категория', 'Цена', 'Остаток', 'Продано'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for product in Product.objects.select_related('category').all():
        sold = (
            OrderItem.objects.filter(product=product, order__status__in=['confirmed', 'assembling', 'shipped'])
            .aggregate(total=Sum('quantity'))['total'] or 0
        )
        ws2.append([
            product.name,
            product.category.name,
            float(product.price),
            product.stock,
            sold,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'sales_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
